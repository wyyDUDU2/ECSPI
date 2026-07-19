# -*- coding: utf-8 -*-
"""
================================================================================
 棉花分类结果 × 统计年鉴 / 参考数据  县市级相关性分析
================================================================================
功能：
  Part 1  分类结果面积  vs  统计年鉴面积   （逐县市，2019/2020/2021 各一图）
  Part 2  分类结果面积  vs  参考数据面积   （逐县市，限定在 aksu/kashi/changji
                                            三个 bounds 内，2019/2020/2021 各一图）

数据：
  - 分类结果 tif（本地）：{region}_M4_RF_CSP_NDVI_{year}*.tif  （可含多块瓦片）
  - 参考数据（GEE asset）：projects/wangyiyao/assets/{year}_Xinjiang
  - 县级矢量（本地）：2024年初县级.shp
  - 统计年鉴（本地 xlsx）：棉花列(第10列, 单位 千公顷)，内页自动识别数据年份

运行环境：服务器  Linux + Python3 + earthengine-api + rasterio + geopandas
  pip install earthengine-api geemap rasterio geopandas openpyxl \
              scikit-learn scipy matplotlib pandas numpy shapely pyproj
  首次使用 GEE 需： earthengine authenticate   （或 ee.Authenticate()）

绘图说明（本版改动）：
  - 全图英文化；标题/标签/刻度/统计文字统一加粗，字体加大。
  - 散点、回归线、1:1 线、坐标轴线均加粗。
  - 拼接图子图编号 (a)(b)(c) / (d)(e)(f) 按行排布，放在左上角统计文字框上方；
    统计文字框整体下移，给编号腾位。
  - 参考数据图（下排）：兵团相关县(BINGTUAN_ALL) 用橙色菱形标记区分（数据值不变，
    仅改标记），图例标 "XPCC county"；上排兵团师仍标 "XPCC division"。

作者备注：
  - 年鉴“棉花”列单位为 千公顷，脚本统一换算为 km^2（×10）。
  - 分类/参考栅格面积按像元真实面积求和换算为 km^2。
  - 相关性(R²、Pearson r)与尺度无关，但同单位便于画 1:1 线。
================================================================================
"""

import ee
import os
import re
import sys
import glob
import json
import math
import warnings
import numpy as np
import pandas as pd

# ---- 依赖检查 ----------------------------------------------------------------
_missing = []
try:
    import rasterio
    import rasterio.mask
except ImportError:
    _missing.append('rasterio')
try:
    import geopandas as gpd
    from shapely.geometry import box, mapping
    from shapely.ops import unary_union
except ImportError:
    _missing.append('geopandas/shapely')
try:
    import openpyxl
except ImportError:
    _missing.append('openpyxl')
try:
    from scipy import stats as sps
except ImportError:
    _missing.append('scipy')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties, findSystemFonts

if _missing:
    print('=' * 60)
    print('缺少以下依赖，请先安装：')
    print('  pip install ' + ' '.join(_missing))
    print('=' * 60)
    sys.exit(1)

warnings.filterwarnings('ignore')


# ============================================================================ #
#                                CONFIG                                        #
# ============================================================================ #
if os.name == 'nt':
    OUTPUT_DIR = r'D:\WYY\ECSPI\期刊'
    FIG_DIR    = r'D:\WYY\figure'
    SHP_PATH   = r'D:\WYY\2024年初县级.shp'
    YB_DIR     = r'D:\WYY'
else:
    OUTPUT_DIR = '/public/ljq/public/wyy'
    FIG_DIR    = '/public/ljq/public/wyy/figure'
    SHP_PATH   = '/public/ljq/public/wyy/2024年初县级.shp'
    YB_DIR     = '/public/ljq/public/wyy'

RESULT_DIR = os.path.join(OUTPUT_DIR, 'correlation_output')
os.makedirs(RESULT_DIR, exist_ok=True)

YEARS = [2019, 2020, 2021]

# 统计年鉴文件（按内页 "(20XX年)" 自动识别其真实数据年份，再与分类年份匹配；
# 若识别失败，则回退用此处的键作为数据年份）。
YEARBOOK_FILES = {
    2019: os.path.join(YB_DIR, '新疆统计年鉴19年.xlsx'),
    2020: os.path.join(YB_DIR, '新疆统计年鉴20年.xlsx'),
    2021: os.path.join(YB_DIR, '新疆统计年鉴21年.xlsx'),
}

# GEE 参考数据
GEE_PROJECT = 'wangyiyao'  # 用于 ee.Initialize(project=...)，按需修改
REF_ASSETS = {
    2019: 'projects/wangyiyao/assets/2019_Xinjiang',
    2020: 'projects/wangyiyao/assets/2020_Xinjiang',
    2021: 'projects/wangyiyao/assets/2021_Xinjiang',
}

# 栅格中“棉花”的取值（请用脚本打印的直方图核对后修改）
COTTON_VALUE_CLASS = 1     # 分类结果 tif 中棉花像元值
COTTON_VALUE_REF   = 1     # 参考数据 asset 中棉花像元值

REF_SCALE     = 30         # reduceRegions 计算参考面积的尺度(米)，过细易超时
GEE_TILESCALE = 4          # reduceRegions tileScale，越大越省内存越慢
EE_SIMPLIFY_DEG = 0.0      # 传给 GEE 的县界简化容差(度)，0=不简化；点过多超时可设 0.001

# 三个分析区域的 bounds（lon_min, lat_min, lon_max, lat_max）
REGION_BOUNDS = {
    'aksu':    (78.0, 40.0, 85.0, 42.5),
    'kashi':   (75.5, 37.5, 80.0, 41.0),
    'changji': (83.0, 43.5, 89.0, 46.0),
}

# 县市 -> 所属分析区域（决定用哪套分类 tif / 哪个 bounds）
COUNTY_REGION = {
    # ---- changji 区（北疆：昌吉/塔城东/克拉玛依/兵团北疆）----
    '乌苏市': 'changji', '胡杨河市': 'changji', '奎屯市': 'changji',
    '克拉玛依区': 'changji', '沙湾市': 'changji', '石河子市': 'changji',
    '玛纳斯县': 'changji', '呼图壁县': 'changji', '昌吉市': 'changji',
    '五家渠市': 'changji', '阜康市': 'changji',
    # ---- aksu 区（阿克苏地区 + 阿拉尔）----
    '库车市': 'aksu', '沙雅县': 'aksu', '新和县': 'aksu', '阿拉尔市': 'aksu',
    '阿克苏市': 'aksu', '阿瓦提县': 'aksu', '温宿县': 'aksu', '柯坪县': 'aksu',
    # ---- kashi 区（喀什/克州 + 图木舒克）----
    '图木舒克市': 'kashi', '巴楚县': 'kashi', '麦盖提县': 'kashi', '莎车县': 'kashi',
    '泽普县': 'kashi', '伽师县': 'kashi', '岳普湖县': 'kashi', '英吉沙县': 'kashi',
    '阿图什县': 'kashi', '喀什市': 'kashi', '疏勒县': 'kashi', '疏附县': 'kashi',
    '阿克陶县': 'kashi',
}
USER_COUNTIES = list(COUNTY_REGION.keys())

# 打星号、效果不好可考虑剔除的县（图中以三角标记，并额外给出剔除后的统计）
OPTIONAL_COUNTIES = {'阿图什县', '喀什市', '疏勒县'}

# 兵团连片团场较多的“地方县”：境内大片棉花在年鉴里被并入“生产建设兵团”汇总、
# 不计入本县棉花列，导致卫星分类(含兵团地块) ≫ 地方年鉴值。
# 这是统计口径错位、非分类误差(它们在“分类 vs 参考”中均贴 1:1)。
# 年鉴图中以方块单独标记，并给出剔除后的 R²。可按需增减。
BINGTUAN_HEAVY = {'乌苏市', '沙湾市', '呼图壁县', '玛纳斯县'}

# 用户名 -> 统计年鉴中的名字（兵团城市年鉴里通常没有，归入“生产建设兵团”，会自动跳过）
YEARBOOK_ALIAS = {
    '阿图什县': '阿图什市',     # 年鉴/行政区实为“阿图什市”
    '克拉玛依区': '克拉玛依市',  # 年鉴仅地级“克拉玛依市”，为近似匹配（偏大）
}

# 兵团城市：统计年鉴常按“生产建设兵团”汇总，缺独立县市值，Part1 会被跳过
BINGTUAN_CITIES = {'胡杨河市', '石河子市', '五家渠市', '阿拉尔市', '图木舒克市'}

QIANHA_TO_KM2 = 10.0       # 1 千公顷 = 1000 公顷 = 10 km^2
REGION_COLOR = {'aksu': '#1f77b4', 'kashi': '#d62728', 'changji': '#2ca02c'}

# 兵团相关县市统一从分析中“直接剔除”：
#   纯兵团市(年鉴无地方值) + 兵团连片混杂县(年鉴口径错位)
BINGTUAN_ALL = BINGTUAN_CITIES | BINGTUAN_HEAVY
# 年鉴与参考分析使用“同一批城市、同样点数”：在用户县市中去掉兵团相关县
ANALYSIS_COUNTIES = [c for c in USER_COUNTIES if c not in BINGTUAN_ALL]

# ---- 绘图样式（字体加大加粗、点线加粗；全英文）----
FS_TITLE   = 26
FS_LABEL   = 22
FS_TICK    = 18
FS_STATS   = 17
FS_LEGEND  = 16
FS_PANEL   = 30          # 子图编号 (a)(b)... 字号
FW_BOLD    = 'bold'      # 统一字重
MARKER_SIZE = 170        # 散点大小
MARKER_EDGE = 1.8        # 散点描边
OLS_LW    = 4.0          # 回归线粗细
LINE11_LW = 2.8          # 1:1 线粗细
AXIS_LW   = 2.0          # 坐标轴/刻度线宽

# 全局加粗（标题/坐标轴标签/刻度/数学文本），保证默认即加粗
matplotlib.rcParams['font.weight']      = 'bold'
matplotlib.rcParams['axes.labelweight'] = 'bold'
matplotlib.rcParams['axes.titleweight'] = 'bold'
matplotlib.rcParams['mathtext.default'] = 'bf'

RUN_PART1 = True
RUN_PART2 = True

# ============================================================================ #
#               兵团“师”级分析配置（让年鉴图凑齐 32 点）                          #
# ============================================================================ #
# 方案：27 个地方县点(地方年鉴) + 5 个兵团师点(兵团年鉴) = 32
#   - 地方县分类面积 = 县界 减去 境内团场镇（只保留地方部分）
#   - 兵团师分类面积 = 该师所有团场镇合并后的范围
RUN_BINGTUAN = True

if os.name == 'nt':
    XZQH_SHP = r'D:\WYY\新疆维吾尔自治区.shp'
    BT_YB_DIR = r'D:\WYY'
else:
    XZQH_SHP = '/public/ljq/public/wyy/新疆维吾尔自治区.shp'
    BT_YB_DIR = '/public/ljq/public/wyy'

# 兵团统计年鉴(.xls)：分师棉花在“棉花”列(实测第16列)，单位千公顷；师块在表下半部
BINGTUAN_YEARBOOK_FILES = {
    2019: os.path.join(BT_YB_DIR, '兵团统计年鉴19年.xls'),
    2020: os.path.join(BT_YB_DIR, '兵团统计年鉴20年.xls'),
    2021: os.path.join(BT_YB_DIR, '兵团统计年鉴21年.xls'),
}

# 参与分析的师（落在三大研究区内）及其所属分类区域
SHI_REGION = {'一师': 'aksu', '三师': 'kashi', '六师': 'changji',
              '七师': 'changji', '八师': 'changji'}

# 师 -> 团场镇 映射（团号 + 镇/农场/牧场名），用于在乡镇 shp 的“乡”字段里匹配。
# shp 中团号多为汉字，匹配时会自动生成“133团/一三三团/一百三十三团”等候选；镇名优先。
SHI_TUAN = {
    '一师': {'tuans': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16],
             'places': ['金银川镇', '新井子镇', '甘泉镇', '永宁镇', '沙河镇', '双城镇',
                        '玛滩镇', '塔门镇', '梨花镇', '昌安镇', '花桥镇', '塔南镇',
                        '幸福镇', '金杨镇', '新开岭镇', '托喀依乡']},
    '三师': {'tuans': [41, 42, 44, 45, 46, 48, 49, 50, 51, 53, 54],
             'places': ['草湖镇', '龙口镇', '永安镇', '前海镇', '永兴镇', '河东镇',
                        '海安镇', '夏河镇', '唐驿镇', '金胡杨镇', '兴安镇', '嘉和镇',
                        '红石榴镇', '兴边镇', '杏花镇', '东风镇', '伽师总场', '红旗农场',
                        '托云牧场', '叶城二牧场', '东风农场']},
    '六师': {'tuans': [101, 102, 103, 105, 106],
             'places': ['青湖镇', '梧桐镇', '蔡家湖镇', '芳草湖', '新湖农场', '军户农场',
                        '共青团农场', '六运湖农场', '土墩子农场', '红旗农场', '奇台农场',
                        '北塔山牧场']},
    '七师': {'tuans': [123, 124, 125, 126, 127, 128, 129, 130, 131, 137],
             'places': ['共青镇', '奎东农场']},
    '八师': {'tuans': [121, 133, 134, 136, 141, 142, 143, 144, 147, 148, 149, 150, 152],
             'places': ['北泉镇', '石河子镇', '石河子总场']},
}
SHI_COLOR = '#ff7f0e'   # 兵团师点颜色（橙）；参考图中的兵团相关县同色


# ============================================================================ #
#                          中文字体（无则回退英文）                              #
# ============================================================================ #
def setup_cjk_font():
    candidates = ['SimHei', 'Microsoft YaHei', 'WenQuanYi Zen Hei', 'WenQuanYi Micro Hei',
                  'Noto Sans CJK SC', 'Source Han Sans SC', 'Source Han Sans CN',
                  'Noto Sans CJK JP', 'PingFang SC', 'Heiti SC', 'Arial Unicode MS']
    installed = {os.path.splitext(os.path.basename(f))[0] for f in findSystemFonts()}
    avail_names = set()
    for f in findSystemFonts():
        try:
            avail_names.add(FontProperties(fname=f).get_name())
        except Exception:
            pass
    for name in candidates:
        if name in avail_names or name in installed:
            matplotlib.rcParams['font.sans-serif'] = [name]
            matplotlib.rcParams['axes.unicode_minus'] = False
            print(f'[font] 使用中文字体: {name}')
            return True
    matplotlib.rcParams['axes.unicode_minus'] = False
    print('[font] 未找到中文字体，图中中文可能显示为方块；建议安装 fonts-wqy-zenhei 或 SimHei')
    return False

# 图全英文，无需依赖中文字体；保留检测仅用于日志/控制台中文标题。
HAS_CJK = setup_cjk_font()


# ============================================================================ #
#                          工具：名称归一化 / 匹配                               #
# ============================================================================ #
def norm_name(s):
    if s is None:
        return ''
    s = re.sub(r'\s+', '', str(s))   # 去掉所有空白：普通/全角/en-space/nbsp/换行
    s = s.lstrip('#').strip()
    return s

def name_core(s):
    """去掉常见行政尾缀，用于模糊匹配。"""
    s = norm_name(s)
    for suf in ['维吾尔自治县', '哈萨克自治县', '蒙古自治县', '回族自治县', '自治县',
                '自治州', '地区', '市辖区', '市', '县', '区']:
        if s.endswith(suf) and len(s) > len(suf):
            return s[:-len(suf)]
    return s


# ============================================================================ #
#                              统计年鉴读取                                     #
# ============================================================================ #
def read_yearbook(path, max_header_row=7):
    """返回 (data_year, {归一化县名: 棉花面积(千公顷) or None})。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    data_year = None
    for r in range(1, min(8, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            m = re.search(r'(20\d{2})\s*年', str(v))
            if m:
                data_year = int(m.group(1))
                break
        if data_year:
            break

    cotton_col = None
    for r in range(1, max_header_row + 1):
        for c in range(1, ws.max_column + 1):
            if norm_name(ws.cell(row=r, column=c).value) == '棉花':
                cotton_col = c
                break
        if cotton_col:
            break
    if cotton_col is None:
        cotton_col = 10  # 经核对样例为第 10 列
        print(f'[yearbook] 警告：未定位到“棉花”表头，回退使用第 {cotton_col} 列')

    data_start = None
    for r in range(1, ws.max_row + 1):
        nm = norm_name(ws.cell(row=r, column=1).value)
        v2 = ws.cell(row=r, column=2).value
        if nm and isinstance(v2, (int, float)):
            data_start = r
            break
    if data_start is None:
        data_start = 8

    out = {}
    for r in range(data_start, ws.max_row + 1):
        nm = norm_name(ws.cell(row=r, column=1).value)
        c2 = norm_name(ws.cell(row=r, column=2).value)
        # 某些年鉴(如2020)在同一 sheet 纵向叠了第二张表(列含甜瓜等)，
        # 读到下一张表的表头("地区"/"农作物"/"播种面积")即停止，避免覆盖棉花值
        if r > data_start and (nm == '地区' or c2 in ('农作物', '播种面积', '农作物播种面积')):
            break
        if not nm:
            continue
        val = ws.cell(row=r, column=cotton_col).value
        if nm not in out:   # 同名只保留首次出现(第一张表)
            out[nm] = float(val) if isinstance(val, (int, float)) else None
    return data_year, cotton_col, out


def load_all_yearbooks():
    """按内页识别的数据年份索引；匹配 YEARS。返回 {year: {county: 千公顷}}。"""
    from openpyxl.utils import get_column_letter
    by_year = {}
    cols_used = {}
    sample_counties = ['阿克苏市', '沙雅县', '伽师县', '昌吉市', '玛纳斯县']
    for tag, path in YEARBOOK_FILES.items():
        if not os.path.exists(path):
            print(f'[yearbook] 缺失文件: {path}')
            continue
        dyear, ccol, table = read_yearbook(path)
        use_year = dyear if dyear in YEARS else tag
        by_year[use_year] = table
        cols_used[use_year] = ccol
        n_ok = sum(v is not None for v in table.values())
        print(f'[yearbook] {os.path.basename(path)} 内页年份={dyear} -> 归为 {use_year}年, '
              f'棉花列={get_column_letter(ccol)}({ccol}列), {n_ok} 个有效县市')
        samp = ', '.join(f'{c}={table.get(norm_name(c))}' for c in sample_counties)
        print(f'           样例(千公顷): {samp}')
    # 列不一致告警（不同文件表头布局不同会导致取错列）
    if len(set(cols_used.values())) > 1:
        print(f'[yearbook] ⚠ 警告：各年棉花列不一致 {cols_used} —— '
              f'某一版表头布局可能不同，请核对偏离的那一版是否取错列！')
    return by_year


def yearbook_area_km2(table, user_name):
    """从年鉴表取某县棉花面积(km^2)；找不到返回 None。"""
    if user_name in BINGTUAN_CITIES:
        return None
    cand = [YEARBOOK_ALIAS.get(user_name, user_name), user_name]
    for c in cand:
        c = norm_name(c)
        if c in table and table[c] is not None:
            return table[c] * QIANHA_TO_KM2
    # 模糊：core 完全相等
    core = name_core(user_name)
    for k, v in table.items():
        if v is not None and name_core(k) == core:
            return v * QIANHA_TO_KM2
    return None


# ============================================================================ #
#                          县界矢量加载 + 字段识别                              #
# ============================================================================ #
def load_counties():
    gdf = gpd.read_file(SHP_PATH)
    if gdf.crs is None:
        gdf.set_crs(epsg=4326, inplace=True)
    gdf4326 = gdf.to_crs(epsg=4326)

    # 自动识别县名字段：选与 USER_COUNTIES 匹配最多的字符串列
    best_field, best_hit = None, -1
    for col in gdf.columns:
        if col == gdf.geometry.name:
            continue
        try:
            vals = gdf[col].astype(str).tolist()
        except Exception:
            continue
        if not any(isinstance(v, str) for v in vals):
            continue
        cores = {name_core(v) for v in vals}
        exacts = {norm_name(v) for v in vals}
        hit = 0
        for uc in USER_COUNTIES:
            if norm_name(uc) in exacts or name_core(uc) in cores \
               or norm_name(YEARBOOK_ALIAS.get(uc, uc)) in exacts:
                hit += 1
        if hit > best_hit:
            best_hit, best_field = hit, col
    print(f'[shp] 县名字段 = {best_field} （匹配 {best_hit}/{len(USER_COUNTIES)}）')

    # 为每个用户县市挑出对应矢量行（4326）
    mapping_geom = {}
    for uc in USER_COUNTIES:
        targets = {norm_name(uc), name_core(uc),
                   norm_name(YEARBOOK_ALIAS.get(uc, uc))}
        sub = gdf4326[gdf4326[best_field].astype(str).apply(
            lambda v: norm_name(v) in targets or name_core(v) in targets)]
        if len(sub) == 0:
            print(f'[shp] 未匹配到县界: {uc}')
            mapping_geom[uc] = None
        else:
            mapping_geom[uc] = unary_union(sub.geometry.values)
    return gdf4326, best_field, mapping_geom


# ============================================================================ #
#                     分类结果(本地 tif) 县级棉花面积                            #
# ============================================================================ #
def find_class_tifs(region, year):
    pat = os.path.join(FIG_DIR, f'{region}_M4_RF_CSP_NDVI_{year}*.tif')
    files = sorted(glob.glob(pat))
    return files


def _pixel_area_km2_grid(shape, transform, crs):
    rows, cols = shape
    if crs is not None and crs.is_projected:
        a = abs(transform.a); e = abs(transform.e)
        return np.full(shape, a * e / 1e6, dtype='float64')
    # 地理坐标(度)：逐行按纬度 cos 修正
    res_x = abs(transform.a); res_y = abs(transform.e)
    r_idx = np.arange(rows)
    lat = transform.f + (r_idx + 0.5) * transform.e   # transform.e<0
    m_lat = 111320.0
    m_lon = 111320.0 * np.cos(np.deg2rad(lat))
    row_area = (res_x * m_lon) * (res_y * m_lat) / 1e6  # km^2, 长度=rows
    return np.repeat(row_area.reshape(-1, 1), cols, axis=1)


def class_area_km2(geom_4326, tif_files, cotton_value):
    """对给定县界(4326)在若干分类瓦片上求棉花面积(km^2)。窗口读取，避免整图入内存。"""
    if geom_4326 is None or geom_4326.is_empty:
        return None
    total = 0.0
    counted = False
    for tp in tif_files:
        with rasterio.open(tp) as ds:
            # 县界投影到栅格 CRS
            try:
                g = gpd.GeoSeries([geom_4326], crs='EPSG:4326').to_crs(ds.crs).iloc[0]
            except Exception:
                g = geom_4326
            # bbox 不相交则跳过
            rb = box(*ds.bounds)
            if not g.intersects(rb):
                continue
            try:
                out, out_tr = rasterio.mask.mask(
                    ds, [mapping(g)], crop=True, all_touched=False,
                    filled=True, nodata=0, indexes=1)
            except Exception:
                continue
            arr = out if out.ndim == 2 else out[0]
            m = (arr == cotton_value)
            if not m.any():
                counted = True
                continue
            pa = _pixel_area_km2_grid(arr.shape, out_tr, ds.crs)
            total += float((pa * m).sum())
            counted = True
    return total if counted else None


def sample_raster_values(tif_files, n=200000):
    """抽样打印栅格取值，便于核对棉花值。"""
    if not tif_files:
        return
    with rasterio.open(tif_files[0]) as ds:
        w = ds.read(1, out_shape=(1, min(ds.height, 1024), min(ds.width, 1024)))
        vals, cnts = np.unique(w[np.isfinite(w)] if np.issubdtype(w.dtype, np.floating) else w,
                               return_counts=True)
        pairs = sorted(zip(vals.tolist(), cnts.tolist()), key=lambda x: -x[1])[:8]
        print(f'  [check] {os.path.basename(tif_files[0])} 取值(降采样) -> {pairs}')


# ============================================================================ #
#                       参考数据(GEE asset) 县级棉花面积                         #
# ============================================================================ #
def init_ee():
    try:
        ee.Initialize(project=GEE_PROJECT)
    except Exception:
        ee.Authenticate()
        ee.Initialize(project=GEE_PROJECT)
    print('[gee] Earth Engine 初始化完成')


def geom_to_ee(geom_4326, simplify_deg=0.0):
    g = geom_4326
    if simplify_deg and simplify_deg > 0:
        g = g.simplify(simplify_deg, preserve_topology=True)
    return ee.Geometry(mapping(g))


def ref_area_km2_for_region(region, year, county_geoms_part2):
    """对该 region 内的县（已与 bounds 求交的几何）一次性 reduceRegions。
    county_geoms_part2: {county: shapely_geom_4326(已∩bounds)}  返回 {county: km^2}。"""
    asset = REF_ASSETS.get(year)
    if asset is None:
        return {}
    bnd = REGION_BOUNDS[region]
    rect = ee.Geometry.Rectangle([bnd[0], bnd[1], bnd[2], bnd[3]])
    ref = ee.Image(asset)
    ref_bin = ref.eq(COTTON_VALUE_REF)
    area_img = ee.Image.pixelArea().updateMask(ref_bin).clip(rect)

    feats = []
    names = []
    for cty, g in county_geoms_part2.items():
        if g is None or g.is_empty:
            continue
        try:
            feats.append(ee.Feature(geom_to_ee(g, EE_SIMPLIFY_DEG), {'NAME': cty}))
            names.append(cty)
        except Exception as e:
            print(f'  [gee] {cty} 几何转换失败: {e}')
    if not feats:
        return {}
    fc = ee.FeatureCollection(feats)
    res = area_img.reduceRegions(collection=fc, reducer=ee.Reducer.sum(),
                                 scale=REF_SCALE, tileScale=GEE_TILESCALE)
    out = {}
    info = res.getInfo()
    for f in info['features']:
        p = f['properties']
        out[p['NAME']] = float(p.get('sum', 0.0)) / 1e6  # m^2 -> km^2
    return out


def ref_value_histogram(year):
    """打印参考 asset 取值直方图，便于核对 COTTON_VALUE_REF。"""
    asset = REF_ASSETS.get(year)
    if not asset:
        return
    try:
        img = ee.Image(asset)
        any_region = REGION_BOUNDS['aksu']
        rect = ee.Geometry.Rectangle(list(any_region))
        h = img.reduceRegion(ee.Reducer.frequencyHistogram(), rect, scale=200,
                             maxPixels=1e9, tileScale=4).getInfo()
        print(f'  [check] {year} 参考asset取值直方图(粗采样): {h}')
    except Exception as e:
        print(f'  [check] 参考直方图获取失败: {e}')


# ============================================================================ #
#                       兵团“师”级处理（年鉴 + 乡镇 shp）                         #
# ============================================================================ #
_CN_DIG = {0: '〇', 1: '一', 2: '二', 3: '三', 4: '四', 5: '五',
           6: '六', 7: '七', 8: '八', 9: '九'}

def _num_digits_cn(n):
    """逐位汉字：133 -> 一三三 ；101 -> 一〇一。"""
    return ''.join(_CN_DIG[int(d)] for d in str(n))

def _num_full_cn(n):
    """标准汉字数：133 -> 一百三十三；21 -> 二十一；101 -> 一百零一。"""
    n = int(n)
    if n < 10:
        return _CN_DIG[n]
    if n < 20:
        return '十' + (_CN_DIG[n % 10] if n % 10 else '')
    if n < 100:
        return _CN_DIG[n // 10] + '十' + (_CN_DIG[n % 10] if n % 10 else '')
    # 百位
    h, rem = n // 100, n % 100
    s = _CN_DIG[h] + '百'
    if rem == 0:
        return s
    if rem < 10:
        return s + '零' + _CN_DIG[rem]
    if rem < 20:
        return s + '一十' + (_CN_DIG[rem % 10] if rem % 10 else '')
    return s + _CN_DIG[rem // 10] + '十' + (_CN_DIG[rem % 10] if rem % 10 else '')

def tuan_candidates(num):
    """某团号的匹配候选（含 阿拉伯/逐位汉字/标准汉字，及 〇<->零 变体）。"""
    forms = {f'{num}团', _num_digits_cn(num) + '团', _num_full_cn(num) + '团'}
    more = set()
    for f in forms:
        more.add(f.replace('〇', '零'))
        more.add(f.replace('零', '〇'))
    return forms | more

_CN2INT = {'零': 0, '〇': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
           '六': 6, '七': 7, '八': 8, '九': 9}

def cn_to_int(s):
    """中文团号转整数。两种写法都支持：
       标准式(含十/百/千)：七十一->71、十一->11、二十->20、一百零三->103
       逐位式(无十百千)：  一三三->133、一零一->101、一三零->130、一团'一'->1"""
    if not s:
        return None
    s = s.replace('〇', '零')
    try:
        if any(u in s for u in '十百千'):
            total, section, num = 0, 0, 0
            for c in s:
                if c in _CN2INT:
                    num = _CN2INT[c]
                elif c == '十':
                    section += (num if num else 1) * 10; num = 0
                elif c == '百':
                    section += (num if num else 1) * 100; num = 0
                elif c == '千':
                    section += (num if num else 1) * 1000; num = 0
                else:
                    return None
            return section + num
        return int(''.join(str(_CN2INT[c]) for c in s))
    except Exception:
        return None


def _read_any_excel(path):
    """稳健读取 .xls/.xlsx（header=None）。旧版 .xls 若 xlrd 失败，自动用
    LibreOffice(soffice) 转 xlsx 再读。返回 DataFrame。"""
    try:
        return pd.ExcelFile(path).parse(0, header=None)
    except Exception:
        pass
    import subprocess, tempfile, glob as _glob
    td = tempfile.mkdtemp()
    try:
        subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx',
                        '--outdir', td, path],
                       check=True, capture_output=True, timeout=180)
        outs = _glob.glob(os.path.join(td, '*.xlsx'))
        if outs:
            return pd.read_excel(outs[0], header=None, engine='openpyxl')
    except Exception as e:
        raise RuntimeError(f'soffice 转换 .xls 失败({e})；可手动转换：'
                           f'soffice --headless --convert-to xlsx "{path}"')
    raise RuntimeError(f'无法读取 {path}')


def read_bingtuan_yearbook(path):
    """读取兵团统计年鉴(.xls)分师棉花。返回 {师标准名(一师..): 棉花千公顷}。"""
    cot = {}
    try:
        ws = _read_any_excel(path)
    except Exception as e:
        print(f'  [bt] 读取失败 {os.path.basename(path)}: {e}')
        return cot
    # 定位棉花列
    cot_col = None
    for r in range(min(8, len(ws))):
        for c in range(ws.shape[1]):
            if norm_name(ws.iat[r, c]) == '棉花':
                cot_col = c
    if cot_col is None:
        cot_col = 15  # 实测第16列(0基=15)
    for r in range(len(ws)):
        nm = norm_name(ws.iat[r, 0])
        m = re.match(r'^([一二三四五六七八九十]+师)', nm)
        if m:
            v = ws.iat[r, cot_col]
            cot[m.group(1)] = float(v) if isinstance(v, (int, float)) and pd.notna(v) else None
    return cot


def load_xiangzhen():
    """加载乡镇 shp，自动识别‘乡’(乡镇名)字段，返回 (gdf4326, name_field)。"""
    gdf = gpd.read_file(XZQH_SHP)
    if gdf.crs is None:
        gdf.set_crs(epsg=4326, inplace=True)
    gdf = gdf.to_crs(epsg=4326)
    # 优先用名为“乡”的字段；否则挑“镇/团/乡/农场”出现最多的字符串列
    name_field = None
    if '乡' in gdf.columns:
        name_field = '乡'
    else:
        best, bestcol = -1, None
        for col in gdf.columns:
            if col == gdf.geometry.name:
                continue
            try:
                vals = gdf[col].astype(str)
            except Exception:
                continue
            hit = vals.str.contains('镇|团|乡|农场|牧场', na=False).sum()
            if hit > best:
                best, bestcol = hit, col
        name_field = bestcol
    print(f'[xiangzhen] 乡镇名字段 = {name_field}  要素数={len(gdf)}')
    return gdf, name_field


def _shi_from_name(nm):
    """直接从乡镇名抽取师号，如 '兵团农六师新湖农场'->'六师'、'第八师'->'八师'。
    仅在名字像兵团(含 兵团/团/农场/牧场/总场/连)时启用，避免误命中。"""
    if not any(t in nm for t in ('兵团', '团', '农场', '牧场', '总场', '连')):
        return None
    m = re.search(r'(?:农|第)?([一二三四五六七八九十]{1,3})师', nm)
    if m:
        return m.group(1) + '师'
    return None


def build_shi_polygons(xz_gdf, name_field):
    """把团场镇匹配到师并合并。优先用乡名里自带的“农X师”；否则回退团号/镇名。
    返回 (shi_poly{师:多边形}, bingtuan_union, report)。"""
    from shapely.ops import unary_union as _uu
    names = xz_gdf[name_field].astype(str).map(norm_name)

    # 各师的镇/农场名候选(回退用)；团号改为“直接解析数字”后查表
    shi_places = {}
    team2shi = {}
    for shi, info in SHI_TUAN.items():
        if shi not in SHI_REGION:
            continue
        shi_places[shi] = {norm_name(p) for p in info['places']}
        for t in info['tuans']:
            team2shi[t] = shi
    shi_keys = {shi: None for shi in SHI_REGION}   # 仅用于报告键

    _team_re = re.compile(r'([一二三四五六七八九十百千零〇]+)团')

    def _match_one(nm):
        # 1) 名字里直接带“农X师/第X师” -> 直接定师
        s = _shi_from_name(nm)
        if s in SHI_REGION:
            return s
        # 2) 解析团号数字 -> 查所属师（仅研究区5个师有表）
        for tok in _team_re.findall(nm):
            v = cn_to_int(tok)
            if v in team2shi:
                return team2shi[v]
        # 3) 镇/农场名整体出现
        for shi, places in shi_places.items():
            for k in places:
                if k and k in nm:
                    return shi
        return None

    assign = {shi: [] for shi in SHI_REGION}
    matched_names = {shi: [] for shi in SHI_REGION}
    used = set()
    bingtuan_like = []
    unmatched_like = []

    for i, nm in names.items():
        if not nm:
            continue
        is_btlike = ('兵团' in nm or any(t in nm for t in ('农场', '牧场', '总场', '连'))
                     or _team_re.search(nm) or _shi_from_name(nm))
        if is_btlike:
            bingtuan_like.append(nm)
        chosen = _match_one(nm)
        if chosen:
            assign[chosen].append(i)
            matched_names[chosen].append(nm)
            used.add(i)
        elif is_btlike:
            unmatched_like.append(nm)

    shi_poly = {}
    for shi, idxs in assign.items():
        if idxs:
            shi_poly[shi] = _uu(xz_gdf.loc[idxs].geometry.values)
    all_idx = sorted(used)
    bingtuan_union = _uu(xz_gdf.loc[all_idx].geometry.values) if all_idx else None

    # 报告
    print('\n[兵团匹配] 各师匹配到的团场镇数：')
    report = {}
    for shi in shi_keys:
        got = matched_names[shi]
        report[shi] = got
        show = '、'.join(got[:8]) + (' …' if len(got) > 8 else '')
        print(f'   {shi}({SHI_REGION[shi]}): {len(got)} 个 -> {show if got else "（无匹配！）"}')
    if unmatched_like:
        print(f'   ⚠ 另有 {len(unmatched_like)} 个疑似兵团乡名未归师，已写入 unmatched 文件')

    # 导出疑似兵团乡名，便于核对/补别名
    try:
        dump = os.path.join(RESULT_DIR, 'bingtuan_xiangzhen_names.txt')
        with open(dump, 'w', encoding='utf-8') as fh:
            fh.write(f'乡镇shp字段“{name_field}”中疑似兵团的乡名（共 {len(bingtuan_like)}）\n')
            fh.write('=' * 50 + '\n')
            for shi in shi_keys:
                fh.write(f'\n【{shi}】匹配 {len(matched_names[shi])} 个:\n')
                for n in matched_names[shi]:
                    fh.write(f'  {n}\n')
            fh.write(f'\n【未归师的疑似兵团乡名】{len(unmatched_like)} 个:\n')
            for n in unmatched_like:
                fh.write(f'  {n}\n')
        print(f'   [save] 疑似兵团乡名清单: {dump}')
    except Exception as e:
        print(f'   [save] 兵团乡名清单写出失败: {e}')

    return shi_poly, bingtuan_union, report


# ============================================================================ #
#                          统计 + 绘图                                          #
# ============================================================================ #
def fit_stats(x, y):
    x = np.asarray(x, float); y = np.asarray(y, float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    n = len(x)
    if n < 2:
        return None
    r = float(np.corrcoef(x, y)[0, 1]) if np.std(x) > 0 and np.std(y) > 0 else float('nan')
    sl, ic, rr, pv, se = sps.linregress(x, y)
    rmse = float(np.sqrt(np.mean((y - (sl * x + ic)) ** 2)))
    bias = float(np.mean(y - x))
    return dict(n=n, r=r, r2=float(rr ** 2), slope=float(sl), intercept=float(ic),
                rmse=rmse, bias=bias, p=float(pv))


def draw_corr(ax, rows, xlabel, ylabel, title, panel=None, mark_bingtuan=False):
    """把一幅相关性散点画到给定 ax 上（全英文、加粗）。
    panel        : 子图编号字母，如 'a'，以 (a) 形式放在左上角统计文字框上方。
    mark_bingtuan: 参考图用——把兵团相关县(BINGTUAN_ALL)画成橙色菱形以作区分，
                   数据值不变，仅改标记。
    返回统计 dict。rows: [{name, region, x, y}]。
    """
    from matplotlib.lines import Line2D
    pts = [r for r in rows if r.get('x') is not None and r.get('y') is not None
           and np.isfinite(r['x']) and np.isfinite(r['y'])]
    if len(pts) < 2:
        ax.set_title(title + ' (insufficient points)', fontsize=FS_TITLE,
                     fontweight=FW_BOLD)
        return None
    xs = [r['x'] for r in pts]; ys = [r['y'] for r in pts]
    st = fit_stats(xs, ys)

    bingtuan_in_ref = False
    for r in pts:
        is_shi = r['region'] == 'bingtuan'
        # 参考图：把兵团相关“县”当作兵团点标成菱形（数据不变，只改标记）
        is_btcounty = (mark_bingtuan and r['name'] in BINGTUAN_ALL)
        if is_shi or is_btcounty:
            col = SHI_COLOR
            marker = 'D'
            size = MARKER_SIZE * 1.25
            bingtuan_in_ref = bingtuan_in_ref or is_btcounty
        else:
            col = REGION_COLOR.get(r['region'], '#555555')
            marker = 'o'
            size = MARKER_SIZE
        ax.scatter(r['x'], r['y'], c=col, s=size, marker=marker,
                   edgecolors='k', linewidths=MARKER_EDGE, alpha=0.88, zorder=3)

    lim = max(max(xs), max(ys)) * 1.08
    ax.plot([0, lim], [0, lim], '--', color='gray', lw=LINE11_LW, zorder=1)
    xx = np.linspace(0, lim, 50)
    ax.plot(xx, st['slope'] * xx + st['intercept'], '-', color='black',
            lw=OLS_LW, zorder=2)
    ax.set_xlim(0, lim); ax.set_ylim(0, lim)
    ax.set_xlabel(xlabel, fontsize=FS_LABEL, fontweight=FW_BOLD)
    ax.set_ylabel(ylabel, fontsize=FS_LABEL, fontweight=FW_BOLD)
    ax.set_title(title, fontsize=FS_TITLE, fontweight=FW_BOLD)
    ax.tick_params(axis='both', labelsize=FS_TICK, width=AXIS_LW, length=7)
    for lab in ax.get_xticklabels() + ax.get_yticklabels():
        lab.set_fontweight(FW_BOLD)
    for sp in ax.spines.values():
        sp.set_linewidth(AXIS_LW)
    ax.set_aspect('equal', adjustable='box')

    # 子图编号 (a)(b)... —— 放在统计文字框上方
    if panel:
        ax.text(0.04, 0.965, f'({panel})', transform=ax.transAxes,
                va='top', ha='left', fontsize=FS_PANEL, fontweight=FW_BOLD)

    sign = '+' if st['intercept'] >= 0 else '-'
    txt = (f"n = {st['n']}\n"
           f"R$^2$ = {st['r2']:.3f}   r = {st['r']:.3f}\n"
           f"y = {st['slope']:.3f}x {sign} {abs(st['intercept']):.2f}\n"
           f"RMSE = {st['rmse']:.1f}   Bias = {st['bias']:+.1f}")
    # 文字框下移：原 y=0.96 -> 0.88，给上方编号腾位
    ax.text(0.04, 0.88, txt, transform=ax.transAxes, va='top', ha='left',
            fontsize=FS_STATS, fontweight=FW_BOLD,
            bbox=dict(boxstyle='round', fc='white', ec='0.5', alpha=0.92, lw=1.6))

    regions_present = {r['region'] for r in pts}
    handles = [Line2D([0], [0], marker='o', color='w', label=rg,
                      markerfacecolor=REGION_COLOR[rg], markeredgecolor='k',
                      markersize=14, markeredgewidth=1.4)
               for rg in ['changji', 'aksu', 'kashi'] if rg in regions_present]
    if 'bingtuan' in regions_present or bingtuan_in_ref:
        lbl = 'XPCC division' if 'bingtuan' in regions_present else 'XPCC county'
        handles.append(Line2D([0], [0], marker='D', color='w', label=lbl,
                              markerfacecolor=SHI_COLOR, markeredgecolor='k',
                              markersize=15, markeredgewidth=1.4))
    handles += [Line2D([0], [0], ls='--', color='gray', lw=LINE11_LW, label='1:1'),
                Line2D([0], [0], ls='-', color='black', lw=OLS_LW, label='OLS')]
    leg = ax.legend(handles=handles, loc='lower right', fontsize=FS_LEGEND,
                    framealpha=0.9)
    for t in leg.get_texts():
        t.set_fontweight(FW_BOLD)
    return st


def save_single(rows, xlabel, ylabel, title, out_png, panel=None, mark_bingtuan=False):
    """单幅图另存(可选)。"""
    fig, ax = plt.subplots(figsize=(8.6, 8.6))
    st = draw_corr(ax, rows, xlabel, ylabel, title, panel=panel,
                   mark_bingtuan=mark_bingtuan)
    fig.tight_layout(); fig.savefig(out_png, dpi=200); plt.close(fig)
    if st:
        print(f'  [plot] 已保存: {out_png}  (n={st["n"]}, R2={st["r2"]:.3f})')
    return st


# ============================================================================ #
#                                  MAIN                                        #
# ============================================================================ #
def main():
    print('\n' + '=' * 60)
    print('Step 0  加载县界与统计年鉴')
    print('=' * 60)
    gdf4326, name_field, county_geom = load_counties()
    yb_tables = load_all_yearbooks()

    # ---- 预计算：每个区域每年分类瓦片，并抽样核对取值 ----
    print('\n' + '=' * 60)
    print('Step 1  分类结果县级面积（本地 tif 分区窗口统计）')
    print('=' * 60)
    region_year_tifs = {}
    for region in REGION_BOUNDS:
        for year in YEARS:
            tifs = find_class_tifs(region, year)
            region_year_tifs[(region, year)] = tifs
            if tifs:
                print(f'  {region} {year}: {len(tifs)} 个瓦片')
                sample_raster_values(tifs)
                # 单块且文件名带非零第二偏移 -> 可能缺配套块
                for tp in tifs:
                    m = re.search(r'-(\d{10})-(\d{10})\.tif$', os.path.basename(tp))
                    if m and len(tifs) == 1 and (m.group(1) != '0000000000'
                                                 or m.group(2) != '0000000000'):
                        sib = os.path.basename(tp).replace(
                            f'-{m.group(1)}-{m.group(2)}.tif', '-0000000000-0000000000.tif')
                        print(f'  ⚠ {region} {year} 仅有偏移块 {m.group(1)}-{m.group(2)}，'
                              f'疑似缺配套块: {sib} —— 该区该年面积可能不完整！')
            else:
                print(f'  {region} {year}: 未找到 tif（pattern={region}_M4_RF_CSP_NDVI_{year}*.tif）')

    # 分类面积：Part1 用完整县界；Part2 用县界∩bounds
    class_full = {y: {} for y in YEARS}   # {year: {county: km2}}
    class_clip = {y: {} for y in YEARS}   # part2
    geom_clip  = {y: {} for y in YEARS}   # {year: {county: shapely ∩bounds}}
    for cty in USER_COUNTIES:
        region = COUNTY_REGION[cty]
        g = county_geom.get(cty)
        rect = box(*REGION_BOUNDS[region])
        g_clip = None if g is None else g.intersection(rect)
        for year in YEARS:
            tifs = region_year_tifs[(region, year)]
            class_full[year][cty] = class_area_km2(g, tifs, COTTON_VALUE_CLASS)
            class_clip[year][cty] = class_area_km2(g_clip, tifs, COTTON_VALUE_CLASS)
            geom_clip[year][cty] = g_clip

    # ---- 参考数据县级面积（GEE） ----
    ref_area = {y: {} for y in YEARS}
    if RUN_PART2:
        print('\n' + '=' * 60)
        print('Step 2  参考数据县级面积（GEE reduceRegions, 限定 bounds）')
        print('=' * 60)
        try:
            init_ee()
            for year in YEARS:
                ref_value_histogram(year)
                for region in REGION_BOUNDS:
                    sub = {c: geom_clip[year][c] for c in USER_COUNTIES
                           if COUNTY_REGION[c] == region and geom_clip[year][c] is not None}
                    if not sub:
                        continue
                    res = ref_area_km2_for_region(region, year, sub)
                    ref_area[year].update(res)
                    print(f'  {region} {year}: 参考面积计算 {len(res)} 个县')
        except Exception as e:
            print(f'[gee] Part2 参考数据计算失败，已跳过 Part2：{e}')
            RUN_PART2_OK = False
        else:
            RUN_PART2_OK = True
    else:
        RUN_PART2_OK = False

    # ============================ Step 2.5  兵团师 ============================ #
    shi_poly, bingtuan_union = {}, None
    shi_class = {y: {} for y in YEARS}
    shi_yb = {y: {} for y in YEARS}
    class_local = {y: {} for y in YEARS}   # 地方县扣除团场后的分类面积
    RUN_BT_OK = False
    if RUN_BINGTUAN:
        print('\n' + '=' * 60)
        print('Step 2.5  兵团师级（乡镇shp按师合并 + 兵团年鉴分师棉花）')
        print('=' * 60)
        try:
            xz, xz_field = load_xiangzhen()
            shi_poly, bingtuan_union, _ = build_shi_polygons(xz, xz_field)
            # 师分类面积（用该师所属区域的分类tif）
            for shi, poly in shi_poly.items():
                reg = SHI_REGION[shi]
                for year in YEARS:
                    shi_class[year][shi] = class_area_km2(
                        poly, region_year_tifs[(reg, year)], COTTON_VALUE_CLASS)
            # 师年鉴棉花
            for year in YEARS:
                bf = BINGTUAN_YEARBOOK_FILES.get(year)
                tab = read_bingtuan_yearbook(bf) if bf and os.path.exists(bf) else {}
                line = []
                for shi in SHI_REGION:
                    v = tab.get(shi)
                    shi_yb[year][shi] = (v * QIANHA_TO_KM2) if v is not None else None
                    line.append(f'{shi}={shi_yb[year][shi]}')
                print(f'   {year} 兵团年鉴(km^2): ' + '  '.join(line))
            # 地方县分类面积 = 县界 扣除 所有团场镇 后
            for cty in USER_COUNTIES:
                g = county_geom.get(cty); reg = COUNTY_REGION[cty]
                if g is None:
                    g_local = None
                elif bingtuan_union is not None:
                    try:
                        g_local = g.difference(bingtuan_union)
                    except Exception:
                        g_local = g
                else:
                    g_local = g
                for year in YEARS:
                    class_local[year][cty] = class_area_km2(
                        g_local, region_year_tifs[(reg, year)], COTTON_VALUE_CLASS)
            RUN_BT_OK = True
        except Exception as e:
            print(f'[bt] 兵团师级处理失败，回退为不含师的方案：{e}')
            RUN_BT_OK = False

    # ============================ Step 3  组装 + 绘图 ============================ #
    print('\n' + '=' * 60)
    print('Step 3  组装结果 + 绘图（年鉴=27地方县+5兵团师=32点；参考=32县）')
    print('=' * 60)
    need_ref = RUN_PART2 and RUN_PART2_OK

    all_records = []
    idx = {}
    for year in YEARS:
        yb = yb_tables.get(year, {})
        for cty in USER_COUNTIES:
            region = COUNTY_REGION[cty]
            yba = yearbook_area_km2(yb, cty) if yb else None
            rec = dict(year=year, kind='county', name=cty, region=region,
                       is_bingtuan_city=(cty in BINGTUAN_CITIES),
                       class_full_km2=class_full[year][cty],
                       class_local_km2=class_local[year].get(cty),
                       class_clip_km2=class_clip[year][cty],
                       yearbook_km2=yba, ref_km2=ref_area[year].get(cty))
            all_records.append(rec)
            idx[(year, cty)] = rec

    # ---- 年鉴行：27 地方县(扣团场) + 5 兵团师 ----
    local_counties = [c for c in USER_COUNTIES if c not in BINGTUAN_CITIES]  # 27
    # 全英文标签
    ylab     = 'Classified cotton area (km$^2$)'
    xlab_yb  = 'Yearbook cotton area (km$^2$)'
    xlab_ref = 'Reference cotton area (km$^2$)'

    rows_yb, rows_ref = {}, {}
    for year in YEARS:
        ry = []
        for c in local_counties:
            ycls = class_local[year].get(c) if RUN_BT_OK else class_full[year][c]
            ry.append(dict(name=c, region=COUNTY_REGION[c],
                           x=idx[(year, c)]['yearbook_km2'], y=ycls))
        if RUN_BT_OK:
            for shi in SHI_REGION:
                ry.append(dict(name=shi, region='bingtuan',
                               x=shi_yb[year].get(shi), y=shi_class[year].get(shi)))
        rows_yb[year] = ry
        # 参考行：32 县
        rows_ref[year] = [dict(name=c, region=COUNTY_REGION[c],
                               x=ref_area[year].get(c), y=class_clip[year][c])
                          for c in USER_COUNTIES]

    # 统计有效点
    def n_valid(rows):
        return sum(1 for r in rows if r['x'] is not None and r['y'] is not None
                   and np.isfinite(r['x']) and np.isfinite(r['y']))
    print('\n各图有效点数：')
    for year in YEARS:
        print(f'   {year}: 年鉴={n_valid(rows_yb[year])}  参考={n_valid(rows_ref[year])}')

    # 输出使用的点名单
    points_txt = os.path.join(RESULT_DIR, 'analysis_points.txt')
    with open(points_txt, 'w', encoding='utf-8') as fh:
        fh.write('年鉴图所用点 = 27 地方县(扣除境内团场镇) + 5 兵团师 = 32\n\n')
        fh.write(f'地方县 {len(local_counties)} 个:\n')
        for rg in ['changji', 'aksu', 'kashi']:
            gs = [c for c in local_counties if COUNTY_REGION[c] == rg]
            fh.write(f'  [{rg}] {len(gs)}: {"、".join(gs)}\n')
        fh.write(f'\n兵团师 {len(SHI_REGION)} 个:\n')
        for shi, reg in SHI_REGION.items():
            fh.write(f'  {shi} ({reg})\n')
        fh.write(f'\n参考图所用点 = 全部 {len(USER_COUNTIES)} 县\n')
    print(f'[save] 点名单: {points_txt}')

    # ---- 拼图：上三幅=年鉴(32), 下三幅=参考(32) ----
    # 子图编号按行：上排 (a)(b)(c)，下排 (d)(e)(f)
    stats_summary = []
    nrows = 2 if need_ref else 1
    panel_letters = [['a', 'b', 'c'], ['d', 'e', 'f']]
    fig, axes = plt.subplots(nrows, 3, figsize=(8.6 * 3, 8.6 * nrows), squeeze=False)
    for j, year in enumerate(YEARS):
        t = f'{year} Classified vs Yearbook'
        st = draw_corr(axes[0, j], rows_yb[year], xlab_yb, ylab, t,
                       panel=panel_letters[0][j])
        if st:
            stats_summary.append(dict(year=year, analysis='class_vs_yearbook', **st))
    if need_ref:
        for j, year in enumerate(YEARS):
            t = f'{year} Classified vs Reference'
            st = draw_corr(axes[1, j], rows_ref[year], xlab_ref, ylab, t,
                           panel=panel_letters[1][j], mark_bingtuan=True)
            if st:
                stats_summary.append(dict(year=year, analysis='class_vs_reference', **st))
    fig.tight_layout(pad=2.4)
    montage = os.path.join(RESULT_DIR, 'corr_montage_yearbook_top_reference_bottom.png')
    fig.savefig(montage, dpi=200)
    plt.close(fig)
    print(f'\n[plot] 拼接图已保存: {montage}')

    # ---- 同时另存 6 幅单图（英文标题；编号沿用拼图）----
    for j, year in enumerate(YEARS):
        if RUN_PART1:
            save_single(rows_yb[year], xlab_yb, ylab,
                        f'{year} Classified vs Yearbook',
                        os.path.join(RESULT_DIR, f'corr_{year}_class_vs_yearbook.png'),
                        panel=panel_letters[0][j])
        if need_ref:
            save_single(rows_ref[year], xlab_ref, ylab,
                        f'{year} Classified vs Reference',
                        os.path.join(RESULT_DIR, f'corr_{year}_class_vs_reference.png'),
                        panel=panel_letters[1][j], mark_bingtuan=True)

    # ---- 保存表格 ----
    # 追加兵团师记录
    if RUN_BT_OK:
        for year in YEARS:
            for shi in SHI_REGION:
                all_records.append(dict(
                    year=year, kind='shi', name=shi, region=SHI_REGION[shi],
                    is_bingtuan_city=False,
                    class_full_km2=None, class_local_km2=None,
                    class_clip_km2=shi_class[year].get(shi),
                    yearbook_km2=shi_yb[year].get(shi),
                    ref_km2=None))
    df = pd.DataFrame(all_records)
    keep = ['year', 'kind', 'region', 'name', 'is_bingtuan_city',
            'class_full_km2', 'class_local_km2', 'class_clip_km2',
            'yearbook_km2', 'ref_km2']
    df = df[[c for c in keep if c in df.columns]]
    csv_path = os.path.join(RESULT_DIR, 'area_table.csv')
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')

    sdf = pd.DataFrame(stats_summary)
    scsv = os.path.join(RESULT_DIR, 'correlation_stats.csv')
    if len(sdf):
        sdf = sdf[['year', 'analysis', 'n', 'r', 'r2',
                   'slope', 'intercept', 'rmse', 'bias', 'p']]
        sdf.to_csv(scsv, index=False, encoding='utf-8-sig')

    # Excel 汇总（两个 sheet）
    try:
        xlsx_path = os.path.join(RESULT_DIR, 'correlation_results.xlsx')
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as w:
            df.to_excel(w, sheet_name='areas', index=False)
            if len(sdf):
                sdf.to_excel(w, sheet_name='correlation_stats', index=False)
        print(f'\n[save] Excel 汇总: {xlsx_path}')
    except Exception as e:
        print(f'[save] Excel 写出失败（忽略）：{e}')

    print(f'[save] 面积表: {csv_path}')
    print(f'[save] 相关性统计: {scsv}')
    print(f'[save] 拼接图(上年鉴/下参考): {montage}')
    print('\n==== 相关性统计汇总 ====')
    if len(sdf):
        with pd.option_context('display.max_rows', None, 'display.width', 160):
            print(sdf.to_string(index=False))
    print('\n全部完成。图与表位于:', RESULT_DIR)


if __name__ == '__main__':
    main()